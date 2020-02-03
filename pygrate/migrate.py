import copy
import os
import argparse
import logging
import shutil
from collections import defaultdict
from functools import partial
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet, Cell

from pygrate.common import SourceAction


LOG = logging.getLogger(__name__)


def read_migration_sheet(path, sheet_name=None):
    """ Simple reader to ingest migration sheets """
    wb = load_workbook(path)
    sheet = wb[sheet_name] if sheet_name else wb.active
    return sheet


class Action:
    def __init__(
        self,
        action,
        source,
        target=None,
        priority=None
    ):
        if not action:
            raise ValueError('Action cannot be none')

        self.action = SourceAction(action)
        self.source = source

        if self.action in (SourceAction.COPY, SourceAction.MOVE) and not target:
            raise ValueError(f'Target needs to be specified for {self.action} on {self.source}')
        self.target = target
        self._target_is_file = bool(
            self.target.suffix) if self.target else False

        self.priority = priority
        self._ignore_sub_folders = []

    @property
    def target_is_file(self):
        return self._target_is_file

    def mark_target_as_file(self):
        self._target_is_file = True
    
    def ignore_sub_folder(self, path):
        if self.source.is_file():
            raise ValueError(f'Source is a file and cannot contain other folders: {self}')

        self._ignore_sub_folders.append(path)

    def __repr__(self):
        res = f'{self.action} {self.source}'
        if self.action in (SourceAction.COPY, SourceAction.MOVE):
            res += f' -> {self.target}'
        return res

    def _delete(self, dry_run):
        if dry_run:
            LOG.info(f'Would delete {self.source}')
        else:
            if self.source.is_file():
                self.source.unlink()
            else:
                shutil.rmtree(str(self.source))

    def _migrate_with_source_name(self, dry_run):
        a = Action(
            self.action,
            self.source,
            self.target / self.source.name,
            self.priority
        )

        if self.source.is_file():
            a.mark_target_as_file()

        a.perform(dry_run=dry_run)

    def _migrate_elements(self, dry_run):
        for entry in self.source.iterdir():
            if entry in self._ignore_sub_folders:
                continue

            a = Action(
                self.action,
                entry,
                self.target / entry.name,
                self.priority
            )

            if a.source.is_dir():
                for path in self._ignore_sub_folders:
                    a.ignore_sub_folder(path)
            
            a.perform(dry_run=dry_run)

    def _migrate(self, func, dry_run):
        if self.target.exists() and not self.target.is_dir():
            raise IOError(f'Target exists: {self}')

        if self.source.is_dir() and self.target_is_file:
            raise IOError(f'Cannot migrate a directory to a file: {self}')

        # is the target an existing directory, change target to
        # target / source.name
        if self.source.is_dir() and self.target.is_dir():
            # migrate all elements in source if names are the same
            if self.source.name.lower() == self.target.name.lower():
                self._migrate_elements(dry_run)

                # clean up if moving things here
                if self.action == SourceAction.MOVE and not dry_run:
                    self.source.rmdir()
            else:
                self._migrate_with_source_name(dry_run)

        elif self.source.is_file() and not self.target_is_file:
            self._migrate_with_source_name(dry_run)

        else:
            target_parent = self.target.parent
            if not target_parent.exists():
                if dry_run:
                    LOG.info(f'Would create directory path: {target_parent}')
                else:
                    # parent does not exist, lets try and create it
                    target_parent.mkdir(parents=True)

            if dry_run:
                LOG.info(
                    f'Would use {func.__name__} to migrate {self.source} -> {self.target}')
            else:
                func(str(self.source), str(self.target))

    def _copy(self, dry_run):
        func = shutil.copy2 if self.source.is_file() else shutil.copytree

        def _ignore_sub_folder_callback(parent, contents):
            ignore = []
            for path in contents:
                if Path(parent + os.path.sep + path) in self._ignore_sub_folders:
                    ignore.append(path)
            return ignore

        if self._ignore_sub_folders:
            LOG.debug(self._ignore_sub_folders)
            func = partial(func, ignore=_ignore_sub_folder_callback)
            func.__name__ = shutil.copytree.__name__

        self._migrate(func, dry_run)

    def _move(self, dry_run):
        self._migrate(shutil.move, dry_run)

    def perform(self, dry_run=False):
        action_msg = 'dry-run' if dry_run else 'perform'
        LOG.info(f'About to {action_msg}: {self}')

        if self.action == SourceAction.DELETE:
            self._delete(dry_run)
        elif self.action == SourceAction.COPY:
            self._copy(dry_run)
        elif self.action == SourceAction.MOVE:
            self._move(dry_run)
        elif self.action == SourceAction.IGNORE:
            LOG.info(f'Ignoring source: {self.source}')
        elif self.action == SourceAction.NOT_DEFINED:
            raise ValueError('Action not defined')
        else:
            raise ValueError(f'Unknown action: {self.action}')

        if not dry_run:
            LOG.info(f'Action performed: {self}')

    __str__ = __repr__


def sheet_to_actions(sheet: Worksheet):
    iter_ = sheet.iter_rows()
    next(iter_)  # skip header

    # collect actions and paths
    actions = {}
    paths = []
    for row in iter_:
        path, _, _, _, action, target, *_ = map(lambda x: x.value, row)

        path = Path(path)
        target = Path(target) if target else None

        if target and not action:
            raise ValueError(f'Target defined without action: {target}')

        if action:
            action_cls = Action(action, path, target, len(path.parents))
            LOG.debug(f'Found action: {action_cls}')
            actions[path] = action_cls
        else:
            paths.append(path)

    # check if all paths are addressed
    for path in paths:
        has_parent_action = any(p in actions for p in path.parents)

        if not has_parent_action:
            LOG.warning(f'{path} has no action')

    return actions


def _prioritize_actions(actions):
    return sorted(actions.values(), key=lambda a: a.priority, reverse=True)


def _convert_encapsulated_actions(actions):
    actions_modified = copy.deepcopy(actions)

    for path, action in actions.items():
        # is action encapsulated?
        for parent in action.source.parents:
            if parent in actions_modified and actions[parent].action is not SourceAction.IGNORE:
                parent_action = actions_modified[parent]

                # if parent action and action are the same, pop it
                if parent_action.action == action.action:
                    LOG.info(f'Removing encapsulated {action} addressed in {parent_action}')
                    actions_modified.pop(path)

                # if action is ignore
                if action.action is SourceAction.IGNORE:
                    if parent_action.action is SourceAction.COPY:
                        LOG.info(f'Adding {action.source} as ignored directory to {parent_action}')
                        parent_action.ignore_sub_folder(path)
                        actions_modified.pop(path)
                    elif parent_action.action is SourceAction.MOVE:
                        LOG.info(f'Converting {action} into delete due to encapsulation in {parent_action}')
                        actions_modified[path] = Action(SourceAction.DELETE, action.source, priority=action.priority)
                
                break  # just work on the first matching parent

    return actions_modified


def perform_actions(actions, dry_run=False):
    actions = _convert_encapsulated_actions(actions)
    actions = _prioritize_actions(actions)
    for action in actions:
        action.perform(dry_run=dry_run)


def dry_run_actions(actions):
    perform_actions(actions, dry_run=True)


def migrate(workbook_path, sheet_name, dry_run=False):
    sheet = read_migration_sheet(workbook_path, sheet_name)
    actions = sheet_to_actions(sheet)
    if dry_run:
        dry_run_actions(actions)
    else:
        perform_actions(actions)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('workbook')
    parser.add_argument('--sheet')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    # configure logging
    logging.basicConfig(level=logging.INFO)

    migrate(args.workbook, args.sheet, args.dry_run)


if __name__ == '__main__':
    main()
